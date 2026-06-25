"""
excel_logger.py
---------------
Manages the master Excel log file (Railway_Files_Log.xlsx).
Appends file name + clickable hyperlink after each file is organised.
Uses openpyxl — fully offline, no Excel installation needed to write,
but Excel must be installed to open the file automatically.
"""

import os
import subprocess
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_BG    = "1A1A2E"   # Deep navy
HEADER_FG    = "E0E0E0"   # Light text
ROW_ODD      = "F0F4FF"   # Soft blue-white
ROW_EVEN     = "FFFFFF"   # White
ACCENT       = "2563EB"   # Railway blue (hyperlink)
BORDER_COLOR = "CBD5E1"   # Subtle border

HEADERS = ["S.No", "Date & Time", "Original File Name", "Category",
           "Final File Name", "Open File"]
COL_WIDTHS = [7, 22, 45, 25, 45, 15]


class ExcelLogger:
    def __init__(self, log_path: str, open_after_update: bool = True):
        """
        Args:
            log_path: Full path to the .xlsx log file.
            open_after_update: If True, open Excel after each update.
        """
        self.log_path = Path(log_path)
        self.open_after_update = open_after_update

    def _make_thin_border(self):
        thin = Side(style="thin", color=BORDER_COLOR)
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _create_workbook(self) -> "Workbook":
        """Create a fresh workbook with styled headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "File Log"

        # Header row
        ws.append(HEADERS)
        header_row = ws[1]
        for i, cell in enumerate(header_row):
            cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._make_thin_border()

        # Column widths
        for i, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header
        ws.freeze_panes = "A2"

        # Row height for header
        ws.row_dimensions[1].height = 30

        # Auto filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

        return wb

    def _style_data_row(self, ws, row_num: int, is_odd: bool, file_path: str):
        """Apply styling to a data row."""
        bg_color = ROW_ODD if is_odd else ROW_EVEN
        border = self._make_thin_border()

        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.font = Font(name="Calibri", size=10)

            if col_idx == len(HEADERS):  # "Open File" hyperlink column
                cell.hyperlink = file_path
                cell.value = "📂 Open"
                cell.font = Font(name="Calibri", size=10, color=ACCENT,
                                 underline="single", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = PatternFill("solid", fgColor=bg_color)

        ws.row_dimensions[row_num].height = 18

    def log(self, original_filename: str, final_filename: str,
            category: str, file_destination: str) -> dict:
        """
        Append a new row to the Excel log.

        Returns:
            dict with success (bool) and error (str) if failed.
        """
        if not OPENPYXL_AVAILABLE:
            return {"success": False,
                    "error": "openpyxl not installed. Run: pip install openpyxl"}

        try:
            # Load or create workbook
            if self.log_path.exists():
                wb = load_workbook(str(self.log_path))
                ws = wb.active
            else:
                self.log_path.parent.mkdir(parents=True, exist_ok=True)
                wb = self._create_workbook()
                ws = wb.active

            # Determine next row
            next_row = ws.max_row + 1
            serial_no = next_row - 1  # Header is row 1
            is_odd = serial_no % 2 != 0
            timestamp = datetime.now().strftime("%d-%b-%Y  %H:%M:%S")

            # Write data
            ws.cell(row=next_row, column=1).value = serial_no
            ws.cell(row=next_row, column=2).value = timestamp
            ws.cell(row=next_row, column=3).value = original_filename
            ws.cell(row=next_row, column=4).value = category
            ws.cell(row=next_row, column=5).value = final_filename

            # Apply styling (includes hyperlink in col 6)
            self._style_data_row(ws, next_row, is_odd, file_destination)

            # Center S.No column
            ws.cell(row=next_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center")

            wb.save(str(self.log_path))

            result = {"success": True, "error": None, "log_path": str(self.log_path)}

            if self.open_after_update:
                self._open_excel()

            return result

        except PermissionError:
            return {"success": False,
                    "error": "Excel file is open. Please close it and try again.",
                    "log_path": str(self.log_path)}
        except Exception as e:
            return {"success": False, "error": str(e), "log_path": str(self.log_path)}

    def _open_excel(self):
        """Open the Excel file using the system default application."""
        try:
            os.startfile(str(self.log_path))
        except Exception:
            # Fallback for non-Windows
            try:
                subprocess.Popen(["start", str(self.log_path)], shell=True)
            except Exception:
                pass

    def read_log(self) -> list:
        """
        Read all entries from the Excel log.
        Returns:
            list of dicts containing log entry fields.
        """
        if not OPENPYXL_AVAILABLE:
            return []
        if not self.log_path.exists():
            return []

        try:
            wb = load_workbook(str(self.log_path), read_only=False)
            ws = wb.active
            entries = []
            
            # The headers are on row 1
            # "S.No", "Date & Time", "Original File Name", "Category", "Final File Name", "Open File"
            for row_idx in range(2, ws.max_row + 1):
                sno = ws.cell(row=row_idx, column=1).value
                dt = ws.cell(row=row_idx, column=2).value
                orig = ws.cell(row=row_idx, column=3).value
                cat = ws.cell(row=row_idx, column=4).value
                final = ws.cell(row=row_idx, column=5).value
                
                open_cell = ws.cell(row=row_idx, column=6)
                hyperlink = open_cell.hyperlink.target if open_cell.hyperlink else ""
                
                # If hyperlink is relative, try to resolve it relative to Excel file directory
                if hyperlink and not os.path.isabs(hyperlink) and not hyperlink.startswith("http"):
                    hyperlink = str(Path(self.log_path).parent / hyperlink)

                # Check for None values and convert to string
                if sno is not None:
                    entries.append({
                        "sno": int(sno) if isinstance(sno, (int, float)) else sno,
                        "datetime": str(dt) if dt is not None else "",
                        "original": str(orig) if orig is not None else "",
                        "category": str(cat) if cat is not None else "",
                        "final": str(final) if final is not None else "",
                        "path": str(hyperlink) if hyperlink else ""
                    })
            wb.close()
            return entries
        except Exception as e:
            import logging
            logging.error(f"Failed to read Excel log: {e}")
            return []

    def write_all_entries(self, entries: list) -> dict:
        """
        Overwrite the entire Excel file with the provided entries.
        entries: list of dicts with keys: 'original', 'final', 'category', 'path', 'datetime' (optional)
        """
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "openpyxl not installed."}

        try:
            self.log_path.parent.mkdir(parents=True, exist_ok=True)
            wb = self._create_workbook()
            ws = wb.active

            for i, entry in enumerate(entries, start=1):
                next_row = i + 1
                is_odd = i % 2 != 0
                
                dt = entry.get("datetime") or datetime.now().strftime("%d-%b-%Y  %H:%M:%S")
                orig = entry.get("original", "")
                cat = entry.get("category", "")
                final = entry.get("final", "")
                path = entry.get("path", "")

                ws.cell(row=next_row, column=1).value = i
                ws.cell(row=next_row, column=2).value = dt
                ws.cell(row=next_row, column=3).value = orig
                ws.cell(row=next_row, column=4).value = cat
                ws.cell(row=next_row, column=5).value = final
                
                # Apply styling & hyperlink
                self._style_data_row(ws, next_row, is_odd, path)
                ws.cell(row=next_row, column=1).alignment = Alignment(
                    horizontal="center", vertical="center")

            wb.save(str(self.log_path))
            return {"success": True, "error": None}
        except PermissionError:
            return {"success": False, "error": "Excel file is open. Please close it and try again."}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def open_log(self):
        """Manually open the Excel log."""
        if self.log_path.exists():
            self._open_excel()

    def get_log_path(self) -> str:
        return str(self.log_path)


# ---------- Quick test ----------

if __name__ == "__main__":
    import tempfile
    with tempfile.TemporaryDirectory() as tmpdir:
        log_path = os.path.join(tmpdir, "Test_Log.xlsx")
        logger = ExcelLogger(log_path, open_after_update=False)

        entries = [
            ("TDR_2024_0056_NIT_Mumbai.pdf", "TDR_2024_0056_NIT_Mumbai.pdf",
             "Tenders", "C:\\Railway Files\\Tenders\\TDR_2024_0056_NIT_Mumbai.pdf"),
            ("Monthly_Report_March.xlsx", "Monthly_Report_March.xlsx",
             "Reports", "C:\\Railway Files\\Reports\\Monthly_Report_March.xlsx"),
            ("Circular_45_Safety.pdf", "Circular_45_Safety.pdf",
             "Circulars", "C:\\Railway Files\\Circulars\\Circular_45_Safety.pdf"),
        ]

        for orig, final, cat, dest in entries:
            r = logger.log(orig, final, cat, dest)
            print(f"Logged '{orig}': {r}")

        print(f"\nLog file created at: {log_path}")
        print(f"File exists: {os.path.exists(log_path)}")
